'''
@title RSAMetGUI
@author: Rebecca Coles
Updated on July 21, 2017
Created on Apr 4, 2017
'''

# Import #######################################################################################
import ntpath
from tkinter import Button, filedialog, PhotoImage, Label, Entry, StringVar
from tkinter.ttk import Separator, Style
from openpyxl import load_workbook
from RSA3DPlot import RSA3DPlot
################################################################################################

class RSAMetGUI(object):
    '''
    GUI for RSA Metrology
    '''
    #Initialize empty sensor data lists
    S00List = []
    S01List = []
    S02List = []
    S10List = []
    S11List = []
    S12List = []
    S20List = []
    S21List = []
    S22List = []

    def __init__(self, master):
        '''
        Constructor
        '''
        self.master = master
        master.title("RSA Meterology")
        
        #Construct GUI
        #Sensor location buttons to open data files for each sensor (CCS Coordinate system)
        #    S22    S12    S02
        #    S21    S11    S01    ^
        #    S20    S10    S00    |+Y
        #                         
        #                <--+X    Ø +Z
        
        #Sensor Buttons
        #REB0
        S00 = Button(master, text="S00",bg = "white", command=lambda:self._loadInputDataFile(S00, RSAMetGUI.S00List, 'S00')) #white
        S00.grid(row=4, column=2, sticky='W')
        
        S01 = Button(master, text="S01",bg = "purple", command=lambda:self._loadInputDataFile(S01, RSAMetGUI.S01List, 'S01')) #purple
        S01.grid(row=3, column=2, sticky='W')

        S02 = Button(master, text="S02",bg = "yellow", command=lambda:self._loadInputDataFile(S02, RSAMetGUI.S02List, 'S02')) #yellow
        S02.grid(row=2, column=2, sticky='W')
        
        #REB1
        S10 = Button(master, text="S10",bg = "cyan", command=lambda:self._loadInputDataFile(S10, RSAMetGUI.S10List, 'S10')) #cyan
        S10.grid(row=4, column=1, sticky='W')

        S11 = Button(master, text="S11",bg = "blue", command=lambda:self._loadInputDataFile(S11, RSAMetGUI.S11List, 'S11')) #blue
        S11.grid(row=3, column=1, sticky='W')

        S12 = Button(master, text="S12",bg = "orange", command=lambda:self._loadInputDataFile(S12, RSAMetGUI.S12List, 'S12')) #orange
        S12.grid(row=2, column=1, sticky='W')
        
        #REB2
        S20 = Button(master, text="S20",bg = "magenta", command=lambda:self._loadInputDataFile(S20, RSAMetGUI.S20List, 'S20')) #magenta
        S20.grid(row=4, column=0, sticky='W')

        S21 = Button(master, text="S21",bg = "green", command=lambda:self._loadInputDataFile(S21, RSAMetGUI.S21List, 'S21')) #green
        S21.grid(row=3, column=0, sticky='W')

        S22 = Button(master, text="S22",bg = "red", command=lambda:self._loadInputDataFile(S22, RSAMetGUI.S22List, 'S22')) #red
        S22.grid(row=2, column=0, sticky='W')
        
        #Add coordinate compass
        self.cordImage = PhotoImage(file="cord.pgm", width=100, height=79)
        Label(image=self.cordImage).grid(row=2, column=3, rowspan=3, sticky='W')
        
        #Grid Spacing
        Label(master, text=" ").grid(row=1, column=0)
        Label(master, text=" ").grid(row=8, column=0)
        Label(master, text=" ").grid(row=15, column=0)
        
        #Grid Separator
        Separator(master, orient="horizontal").grid(row=5, column=0, columnspan=4, sticky='ew')
        Separator(master, orient="horizontal").grid(row=13, column=0, columnspan=4, sticky='ew')
        Style(master).configure("TSeparator", background="black")
        
        #Labels
        Label(master, text="Load Sensor Metrology Files (.xlsx)").grid(row=0, column=0, columnspan=2, sticky='W')
        Label(master, text="Raft Plane Equations (optional)").grid(row=6, column=0, columnspan=2, sticky='W')
        Label(master, text="Ex:    53.0234 + 0.0010629 x + 0.00322188 y").grid(row=7, column=0, columnspan=2, sticky='W')
        Label(master, text="Process and Plot Data").grid(row=14, column=0, columnspan=2, sticky='W')
        
        #Datum Plane Text Box
        datumPlaneEqn = StringVar()
        Label(master, text="Enter Datum Plane Equation (must also provide Raft Fit Equation)").grid(row=9, column=0, columnspan=2, sticky='W')
        Entry(master, textvariable=StringVar(), width=40).grid(row=10, column=0, columnspan=3, sticky='W')
        
        #Raft Fit Text Box
        raftFitEqn = StringVar()
        Label(master, text="Enter Raft Fit Equation (must also provide Datum Plane Equation)").grid(row=11, column=0, columnspan=2, sticky='W')
        Entry(master, textvariable=raftFitEqn, width=40).grid(row=12, column=0, columnspan=3, sticky='W')
        
        #Plot RSA for manually selected sensor positions
        manualPosition = Button(master, text = "Plot virtual RSA for manually selected sensor positions", 
                                command=lambda:RSA3DPlot().createVirtualRSA(RSAMetGUI.S00List, RSAMetGUI.S01List, RSAMetGUI.S02List, 
                                                                            RSAMetGUI.S10List, RSAMetGUI.S11List, RSAMetGUI.S12List, 
                                                                            RSAMetGUI.S20List, RSAMetGUI.S21List, RSAMetGUI.S22List, 0, datumPlaneEqn, raftFitEqn))
        manualPosition.grid(row=16, column=0, columnspan=3, sticky='W')
        
        #Plot RSA for optimal sensor positions
        
        
    def _loadInputDataFile(self, sensorButton, sensorList, sensorButtonLabel):
        '''
        Load metrology data from input file using openpyxl library
        '''
        #open file
        fileName = self._openFile()
        #load workbook
        self.wb = load_workbook(fileName, read_only=True)
        #load Sheet1
        self.ws = self.wb.get_sheet_by_name('Sheet1')
        #create a python list of the data by iterating over all of the Sheet1 data
        del sensorList[:] #empty list in case the user is changing a previously loaded file
        sensorList.append(list(self._iter_rows(self.ws)))
        #update the button text to show the filename
        sensorButton.config(text = sensorButtonLabel + ': ' + self._path_leaf(fileName))

        
    def _openFile(self):
        '''
        Create open file dialogue box
        '''
        return filedialog.askopenfilename()
    
    def _iter_rows(self, ws):
        '''
        iterate through xlsx list
        '''
        for row in self.ws.iter_rows(row_offset=1):
            yield [cell.value for cell in row]
            
    def _path_leaf(self, path):
        '''
        get filename from full path
        '''
        head, tail = ntpath.split(path)
        return tail or ntpath.basename(head)